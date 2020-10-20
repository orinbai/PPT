from openpyxl import Workbook
import json
import pandas as pd
import numpy as np
# openpyxl 不会读取所有的excel工作表内容，
# 所以，不能只改部分数字或内容


class GetData:
    def __init__(self, path, filename):
        self.basePath = path
        self.fname = filename
        self.wb = Workbook()
        self.signHash = {
            "基础表2019": "Region",
            "2020潜客量": "大区",
            "2020发票": "大区"
        }

    def startSign(self, baseSheet="基础表2019"):
        # 找到基础表中最基础数据的开始段
        # 该数据最基础的数据是各4S店的数据
        # 根据当前的特征是从"Region"开始
        # 因为是以4S为单位汇报数据，所以
        # E列必须符合4S的ID规则，现在看起
        # 来是"SA+数字"
        # 如果将来修改规则，需要更新这部分的规则
        upperLeftText = self.signHash[baseSheet]
        validStart = "SA"
        


    
# class MonRep:
#     def __init__(self, df, startCELL="A1"):
#         self.data = df
#         self.wb = Workbook()
#         self.sPOS = startCELL
#         self.posLOOP = self.charLOOP()

#     def charLOOP(self):
#         return dict(enumerate(range(ord("A"), ord("Z")+1)))

#     def header(self, header, sheetIDX=0):
#         # header 使用json传值，可以套嵌
#         # 使用 [起始行, 列名, 行数, 列数]
#         # 比如: [1, "A", 1, 3, "Test"] 对应 A1～C1合并，值为Test
#         #           [1, "A", 2, 3, "Test"] 对应A1 ~ C2 合并，值为Test
#         #          [1, "A", 1, 1, "Test"] 那就是单独的A1，值为Test
#         for c in header:
#             ws = self.wb.active(sheetIDX)
#             if c[2] + c[3] > 2:
#                 ws.merge_cells(self.cellJoint(c))
#                 ws['%s%d' % (c[1], c[0])].value = c[4]

#     def cellJoint(arrList):
#         string1 = "%s%d" % (arrList[1], arrList[0])
#         string2 = "%s%d" % (
#             chr(ord(arrList[1] + arrList[3] - 1)),
#             arrList[2] - arrList[0] + 1
#             )
#         return "%s:%s" % (string1, string2)

#     def saveFile(self, filename="test.xlsx"):
#         self.wb.save(filename)
